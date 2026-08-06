#!/usr/bin/env python3
"""Convierte las planillas de índice a un CSV listo para importar.

Las planillas cambiaron de formato a lo largo de los años, así que en vez de
asumir posiciones fijas se detecta el encabezado y se mapean las columnas por
su nombre. Dos familias conocidas:

  2009-2013 (.xls)   Dcto. Nº | FECHA | CONCEPTO | [MONTO] | EXPTE. Nº | [ORIGEN]
                     La fecha es un número de serie de Excel y el expediente
                     viene entero en un campo: "17701-DEV-12".

  2025 (.xlsx)       Encabezado de dos niveles, con el expediente repartido
                     en TIPO / NUMERO / AÑO y la fecha como texto.

Nada se descarta: cada fila guarda en `origen` todas las celdas tal como
vinieron, indexadas por el nombre de su columna, junto con la hoja y la fila
de procedencia. Las columnas normalizadas son una vista derivada para poder
buscar y filtrar.

    python scripts/importar_indice.py "DGSTO. MPAL. 2.013.xls"
    python scripts/importar_indice.py planilla.xlsx --anio 2024 -o salida.csv
    python scripts/importar_indice.py *.xls -o todos.csv
"""

import argparse
import csv
import json
import re
import sys
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path

CAMPOS = ["tipo", "numero", "anio", "fecha", "indice",
          "exp_tipo", "exp_numero", "exp_anio", "origen"]

# Palabras del nombre de la hoja que identifican el tipo de norma. Se prueban
# de más específica a más general: "resolucion personal" antes que
# "resolucion", que si no se la comería.
TIPOS_HOJA = (
    ("resolucion personal", "resolucion_personal"),
    ("resoluciones", "resolucion"),
    ("resolucion", "resolucion"),
    ("decretos", "decreto"),
    ("decreto", "decreto"),
    ("ordenanzas", "ordenanza"),
    ("ordenanza", "ordenanza"),
    ("disposiciones", "disposicion"),
    ("disposicion", "disposicion"),
)

# Nombre de columna normalizado -> campo. La clave se busca como subcadena,
# así que "dcto n" cubre "Dcto. Nº" y "Dcto. Nº " con espacio al final.
COLUMNAS = (
    ("expediente tipo", "exp_tipo"),
    ("expediente numero", "exp_numero"),
    ("expediente ano", "exp_anio"),
    ("expediente", "expediente"),   # el campo entero, formato viejo
    ("expte", "expediente"),
    ("dcto n", "numero"),
    ("resol n", "numero"),
    ("numero", "numero"),
    ("fecha", "fecha"),
    ("concepto", "indice"),
    ("indice", "indice"),
    ("tramite", "tramite"),
    ("monto", "monto"),
    ("origen", "origen_norma"),
)

# Expediente del formato viejo: número, iniciales de la repartición, y año de
# dos o cuatro dígitos. Ej: 17701-DEV-12, 505895-M-06, 7462-L-1964.
EXPEDIENTE_VIEJO = re.compile(r"(\d+)\s*-\s*([A-Za-zÁÉÍÓÚÑ.]+)\s*-\s*(\d{2,4})")

# Marcas de que no hay expediente.
SIN_EXPEDIENTE = ("s/expte", "sexpte", "sinexpte", "snexpte")

FORMATOS_FECHA = ("%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y")

# Origen del calendario de Excel. El de 1900 arrastra el bug del año bisiesto
# inexistente, que se compensa restando dos días.
EPOCA_EXCEL = datetime(1899, 12, 30)


def normalizar(texto) -> str:
    sin_tildes = unicodedata.normalize("NFKD", str(texto))
    limpio = "".join(c for c in sin_tildes if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", " ", limpio.lower()).strip()


def crudo(valor) -> str:
    """El valor tal como vino de la celda, sin interpretar."""
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        return valor.isoformat()
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def entero(valor):
    if valor is None or valor == "":
        return None
    if isinstance(valor, bool):
        return None
    if isinstance(valor, (int, float)):
        return int(valor)
    m = re.search(r"\d+", str(valor).replace(".", ""))
    return int(m.group()) if m else None


def anio_completo(dos_digitos: int, tope: int) -> int:
    """Expande un año de dos dígitos sin inventar fechas futuras."""
    siglo_21 = 2000 + dos_digitos
    return siglo_21 if siglo_21 <= tope else 1900 + dos_digitos


def parsear_fecha(valor, anio_norma: int):
    """Interpreta la fecha en cualquiera de las formas en que aparece.

    Tolera dos errores de tipeo inequívocos del formato nuevo: el espacio
    antes de los dos puntos ('12 :15:28') y la barra faltante ('16/092025').
    """
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor.isoformat()

    # Número de serie de Excel, usado en las planillas viejas.
    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        if 1 <= valor <= 100_000:
            return (EPOCA_EXCEL + timedelta(days=float(valor))).isoformat()
        return None

    texto = re.sub(r"\s*:\s*", ":", str(valor).strip())
    texto = re.sub(r"\s+", " ", texto)

    candidatos = [texto]
    faltante = re.match(r"^(\d{1,2})/(\d{2})(\d{4})\b(.*)$", texto)
    if faltante:
        d, m, a, resto = faltante.groups()
        candidatos.append(f"{d}/{m}/{a}{resto}")

    for candidato in candidatos:
        for formato in FORMATOS_FECHA:
            try:
                return datetime.strptime(candidato, formato).isoformat()
            except ValueError:
                continue
    return None


def parsear_expediente(texto: str, anio_norma: int):
    """Separa el expediente del formato viejo en sus tres partes.

    Algunas celdas encadenan varios ('210334-M-90 y Ac. Nº 17591-M-08'); se
    toma el primero, que es el principal. El texto completo queda en `origen`.
    """
    if not texto:
        return None, None, None

    if normalizar(texto).replace(" ", "") in SIN_EXPEDIENTE:
        return None, None, None

    m = EXPEDIENTE_VIEJO.search(texto)
    if not m:
        return None, entero(texto), None

    numero, iniciales, anio = m.groups()
    a = int(anio)
    if len(anio) <= 2:
        a = anio_completo(a, anio_norma)

    return iniciales.strip(".").upper() or None, int(numero), a


def tipo_de_hoja(nombre: str):
    n = normalizar(nombre)
    for clave, tipo in TIPOS_HOJA:
        if clave in n:
            return tipo
    return None


def anio_del_nombre(nombre: str):
    """Los años vienen escritos como '2.013', así que se quitan los puntos."""
    m = re.search(r"(?:19|20)\d{2}", str(nombre).replace(".", ""))
    return int(m.group()) if m else None


# --- lectura de planillas -------------------------------------------------

def filas_de_planilla(ruta: Path):
    """Devuelve (nombre_hoja, filas) para cada hoja, sea .xls o .xlsx."""
    if ruta.suffix.lower() == ".xls":
        import xlrd
        wb = xlrd.open_workbook(str(ruta))
        for hoja in wb.sheet_names():
            ws = wb.sheet_by_name(hoja)
            yield hoja, [
                [ws.cell_value(r, c) for c in range(ws.ncols)]
                for r in range(ws.nrows)
            ]
    else:
        import openpyxl
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        for hoja in wb.sheetnames:
            yield hoja, [list(f) for f in wb[hoja].iter_rows(values_only=True)]


def ubicar_encabezado(filas, limite: int = 8):
    """Encuentra la fila de encabezado y compone los nombres de columna.

    La fila de encabezado es la primera que menciona una fecha, presente en
    todos los formatos. Si la fila de arriba trae rótulos que agrupan
    columnas —el caso del formato nuevo, con EXPEDIENTE abarcando tres— se
    los antepone, porque sin eso las dos columnas llamadas NUMERO serían
    indistinguibles.
    """
    for i, fila in enumerate(filas[:limite]):
        nombres = [normalizar(c) if c is not None else "" for c in fila]
        if not any(n == "fecha" for n in nombres):
            continue

        if i > 0:
            grupos, actual = [], ""
            for celda in filas[i - 1]:
                texto = normalizar(celda) if celda is not None else ""
                # Los rótulos de grupo abarcan varias columnas y solo aparecen
                # sobre la primera, así que se arrastran hacia la derecha.
                if texto:
                    actual = texto
                grupos.append(actual)

            if any(g and g not in nombres for g in grupos):
                nombres = [
                    f"{g} {n}".strip() if g and n else n
                    for g, n in zip(grupos + [""] * len(nombres), nombres)
                ]

        return i, nombres

    return None, None


def mapear_columnas(nombres):
    """Nombre de columna -> campo, por coincidencia de subcadena."""
    mapa = {}
    for indice, nombre in enumerate(nombres):
        if not nombre:
            continue
        for clave, campo in COLUMNAS:
            if clave in nombre:
                # La primera columna que reclama un campo se lo queda: los
                # encabezados están ordenados de más específico a más general.
                mapa.setdefault(campo, indice)
                break
    return mapa


def leer_hoja(nombre_hoja, filas, tipo, anio, avisos):
    fila_enc, nombres = ubicar_encabezado(filas)
    if fila_enc is None:
        avisos.append(f"Hoja '{nombre_hoja}': no se encontró el encabezado, se omite")
        return []

    mapa = mapear_columnas(nombres)

    # La columna del texto no siempre se llama igual ni se llama: en las
    # resoluciones de personal el asunto está bajo TRAMITE, y en los decretos
    # de 2025 esa columna directamente no tiene encabezado.
    if "indice" not in mapa:
        if "tramite" in mapa:
            mapa["indice"] = mapa.pop("tramite")
        else:
            usadas = set(mapa.values())
            candidatas = [
                i for i in range(len(nombres))
                if i not in usadas and any(
                    len(crudo(f[i])) > 20 for f in filas[fila_enc + 1:fila_enc + 40]
                    if i < len(f)
                )
            ]
            if candidatas:
                mapa["indice"] = candidatas[0]
                avisos.append(
                    f"Hoja '{nombre_hoja}': la columna {candidatas[0] + 1} no tiene "
                    f"encabezado y se toma como texto del índice"
                )

    for obligatorio in ("numero", "indice"):
        if obligatorio not in mapa:
            avisos.append(
                f"Hoja '{nombre_hoja}': falta la columna de {obligatorio} "
                f"(encabezado leído: {[n for n in nombres if n]}), se omite"
            )
            return []

    def celda(fila, campo):
        i = mapa.get(campo)
        return fila[i] if i is not None and i < len(fila) else None

    resultado = []

    for numero_fila, fila in enumerate(filas[fila_enc + 1:], start=fila_enc + 2):
        if not any(c is not None and str(c).strip() for c in fila):
            continue

        n = entero(celda(fila, "numero"))
        indice = crudo(celda(fila, "indice"))
        tramite = crudo(celda(fila, "tramite"))

        # En las planillas viejas el asunto queda repartido entre CONCEPTO y
        # TRÁMITE, y por separado ninguno se entiende: "Sr. Masotto Carlos" /
        # "Rect. Art. 1º Resol. 707-12". Se unen para que la búsqueda por
        # texto encuentre ambos.
        texto = " — ".join(p for p in (indice, tramite) if p)

        if n is None and not texto:
            continue

        # Todas las celdas, con el nombre de su columna, sin interpretar.
        original = {
            (nombres[i] or f"col{i}"): crudo(v)
            for i, v in enumerate(fila)
            if crudo(v)
        }
        original["hoja"] = nombre_hoja
        original["fila"] = numero_fila

        expediente = crudo(celda(fila, "expediente"))
        if expediente:
            exp_t, exp_n, exp_a = parsear_expediente(expediente, anio)
        else:
            exp_t = crudo(celda(fila, "exp_tipo")) or None
            exp_t = re.sub(r"[^A-Za-z]", "", exp_t).upper() or None if exp_t else None
            exp_n = entero(celda(fila, "exp_numero"))
            exp_a = entero(celda(fila, "exp_anio"))
            if exp_a is not None and 0 <= exp_a <= 99:
                exp_a = anio_completo(exp_a, anio)
            if exp_a is not None and not (1900 <= exp_a <= anio):
                exp_a = None

        normalizado = {
            "numero": n,
            "fecha": parsear_fecha(celda(fila, "fecha"), anio),
            "exp_tipo": exp_t,
            "exp_numero": exp_n,
            "exp_anio": exp_a,
        }

        original["sin_normalizar"] = sorted(
            campo for campo, valor in normalizado.items()
            if valor is None and any(
                k for k in original
                if campo.split("_")[-1] in k and k not in ("hoja", "fila")
            )
        )

        resultado.append({
            "tipo": tipo,
            "numero": "" if n is None else n,
            "anio": anio,
            "fecha": normalizado["fecha"] or "",
            "indice": texto,
            "exp_tipo": exp_t or "",
            "exp_numero": "" if exp_n is None else exp_n,
            "exp_anio": "" if exp_a is None else exp_a,
            "origen": json.dumps(original, ensure_ascii=False),
        })

    return resultado


def procesar(ruta: Path, anio_forzado, avisos):
    anio = anio_forzado or anio_del_nombre(ruta.name)
    if not anio:
        avisos.append(f"{ruta.name}: no se pudo deducir el año, se omite")
        return []

    todas = []
    for nombre_hoja, filas in filas_de_planilla(ruta):
        tipo = tipo_de_hoja(nombre_hoja)
        if not tipo:
            avisos.append(f"Hoja '{nombre_hoja}': tipo de norma no reconocido, se omite")
            continue

        # El año del nombre de la hoja manda sobre el del archivo: hay
        # planillas que arrastran hojas de otro ejercicio.
        anio_hoja = anio_del_nombre(nombre_hoja) or anio

        filas_hoja = leer_hoja(nombre_hoja, filas, tipo, anio_hoja, avisos)
        todas.extend(filas_hoja)
        print(f"    {nombre_hoja:28} {tipo:20} {len(filas_hoja):5} normas")

    return todas


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("planillas", type=Path, nargs="+")
    ap.add_argument("--anio", type=int,
                    help="año de las normas (por defecto se toma del nombre)")
    ap.add_argument("-o", "--salida", type=Path,
                    help="CSV de salida (por defecto, junto a la planilla)")
    args = ap.parse_args()

    todas, avisos = [], []
    for ruta in args.planillas:
        if not ruta.is_file():
            sys.exit(f"No existe el archivo {ruta}")
        print(f"\n  {ruta.name}")
        todas.extend(procesar(ruta, args.anio, avisos))

    if not todas:
        sys.exit("\nNo se leyó ninguna fila.")

    salida = args.salida or args.planillas[0].with_suffix(".csv")
    with open(salida, "w", newline="", encoding="utf-8") as f:
        escritor = csv.DictWriter(f, fieldnames=CAMPOS)
        escritor.writeheader()
        escritor.writerows(todas)

    print(f"\n  {len(todas)} normas -> {salida}")

    sin_fecha = sum(1 for f in todas if not f["fecha"])
    sin_exp = sum(1 for f in todas if not f["exp_numero"])
    print(f"  {sin_fecha} sin fecha normalizada, {sin_exp} sin expediente")
    print("  (el valor de origen se conserva en la columna `origen`)")

    if avisos:
        print(f"\n  {len(avisos)} avisos:")
        for a in avisos[:20]:
            print(f"    {a}")
        if len(avisos) > 20:
            print(f"    ... y {len(avisos) - 20} más")


if __name__ == "__main__":
    main()
